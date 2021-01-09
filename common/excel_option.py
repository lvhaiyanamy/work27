"""
============================
Author:蓝色水汀
Time:2020/12/11  10:14
E-mail:826926575@qq.com
Company:陕西伟业医药有限公司
============================
"""
import openpyxl


class ReadExcel:
    def __init__(self, file_name, sheet_name):
        """
        :param file_name:excel文件名
        :param sheet_name:表单名
        """
        self.file_name = file_name
        self.sheet_name = sheet_name

    def read_excel(self):
        """
        :return:返回用例列表
        """
        # 加载excel工作簿
        workbook = openpyxl.load_workbook(self.file_name)
        # 打开该工作簿的表单
        sh = workbook[self.sheet_name]
        # 按行读取Excel数据，且转换为列表形式存储
        datas = list(sh.rows)
        # 读取第一行数据
        title = [i.value for i in datas[0]]
        # 将聚合的title和data转换为字典，存储在列表中，返回得到的列表
        cases = []
        for item in datas[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        return cases

    def write_excel(self, row, column, value, font):
        """
        :param row:写入的行
        :param column:写入的列
        :param value:写入的值
        :param font:写入的字体格式
        """
        workbook = openpyxl.load_workbook(self.file_name)
        sh = workbook[self.sheet_name]
        sh.cell(row=row, column=column, value=value).font = font
        workbook.save(self.file_name)
